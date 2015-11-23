import os
import os.path
import collections
from glob import glob

from openpyxl import load_workbook, Workbook

def strip_csv_line(l):
    return ','.join(list(map(str.strip, l.strip().split(','))))

def compare_csv(a, b):
    a = os.path.abspath(a)
    b = os.path.abspath(b)

    print('Reading files:')
    print('\t', a)
    print('\t', b)
    
    aUnique = []
    bUnique = []
    intersection = []

    with open(a, mode='r') as fileA, open(b, mode='r') as fileB:
        a = [strip_csv_line(i) for i in fileA.readlines()]
        b = [strip_csv_line(i) for i in fileB.readlines()]
        intersection = list(set(a).intersection(b))
        
        aUnique = [i for i in a if i not in intersection]
        bUnique = [i for i in b if i not in intersection]
        
    return aUnique, bUnique, intersection
    
def remove_duplicates_csv(file, output):
    """ Removes duplicate lines in each file.
    """
    lines = []
    duplicates = 0

    file = os.path.abspath(file)
    output = os.path.abspath(output)
    
    print('Reading file', file)
    
    # OPTIMIZE: this can probably run a lot faster
    with open(file, mode='r') as f:
        for line in f:
            # since it's supposed to be CSV, let's do some simple parsing
            # trim trailing whitespace, break it into an array
            csv = line.strip().split(',')
            # trim each item
            csv = [i.strip() for i in csv]
            # put it back together with no extra spaces
            csv = ','.join(csv)
            if csv in lines:
                duplicates += 1
                continue
            else:
                lines.append(csv)

    print('Removed', duplicates, 'duplicates.')
    print('Writing trimmed file to', output)
    with open(output, mode='w') as f:
        f.write('\n'.join(lines))

def concat_files_csv(files, output, keepFirstHeader=False):
    """ Concat each csv file in files.
    """
    print('Found', len(files), 'CSV files.')
    
    text = []
    l = 0
    for f in files:
        with open(f, mode='r') as csv:
            lines = csv.readlines()
            l += len(lines)
            print("There are", len(lines), 'lines')
            
            if keepFirstHeader and len(text) == 0:
                text.append(lines[0])

            if keepFirstHeader:
                text.extend(lines[1:])
            else:
                text.extend(lines)

        # make sure the file ends with a newline
        if not text[-1].endswith('\n'):
            text[-1] = text[-1] + '\n'

    print('Saving concatenated file to:', output)
    print('\tThere are', l, 'lines.')
    with open(output, mode='w') as o:
        o.write(''.join(text))
    
def concat_files_excel(files, output, keepFirstHeader=False):
    """Concats all excel files in files.
        keepFirstHeader - this will keep the first line of only the first file.
    """
    output = os.path.abspath(output)
    print('Concatenating', len(files), 'Excel files.')

    rows = []
    
    for f in files:
        wb = load_workbook(f)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for i, row in enumerate(sheet.rows):
            if keepFirstHeader and i == 0 and len(rows) > 1:
                continue
            rows.append(row)
        wb.save(f)

    wb = Workbook()
    if len(wb.get_sheet_names()) == 0:
        sheet = wb.create_sheet('Sheet1')
    else:
        sheet = wb.active
    
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            # add 1 to index because cell indices are not zero-based
            sheet.cell(row=i+1, column=j+1).value = cell.value
            
    print('Saving concatenated file to:', output)
    wb.save(output)

def summarize_cells(files, cellDict, log=False):
    """ Return a dict of {cellName : [list of values from each file]}.
      cellDict - dict of {cellName (human readable) : (column, row)}
        cellName should be something human readable, for your sanity.
      returns a results dict that looks like cellDict but where the cell
        coords are replaced with a list of all the values in each file.
    """
    result = collections.OrderedDict()
    print('Summarizing', len(files), 'files.')
    
    for f in files:
        if True:
            print('\tlooking at', os.path.basename(f))
        wb = load_workbook(f, read_only=True)
        sheet = wb.active
        for cellName, coord in cellDict.items():
            v = sheet.cell(column=coord[0], row=coord[1]).value
            result.setdefault(cellName, []).append(v)
            if log:
                 print('\t\tfound', v)

    return result

def get_cell_range(files, firstCell, lastCell, outFile, prune=False):
    """ Create an excel file with a list of cell values from each file.
    """

    out = []
    
    for f in files:
        wb = load_workbook(f, read_only=True)
        sheet = wb.active
        cells = sheet[firstCell:lastCell]
        for row in cells:
            out.append([cell.value for cell in row])

    wb = Workbook()
    sheet = wb.active

    if prune:
        # only keep rows where there is at least one cell that is not None
        # (can probably be done a tad faster and more space efficient)
        out = [row for row in out
               if len([cell for cell in row if cell is not None]) > 0]
        
    for row, r in enumerate(out):
        for col, cell in enumerate(r):
            sheet.cell(column=col+1, row=row+1).value = cell

    print('Saving to', os.path.abspath(outFile))
    wb.save(outFile)

def create_summary_sheet(globPath, cellDict, outFile):
    files = glob(globPath)
    result = summarize_cells(files, cellDict)

    result = [[k] + v for k, v in result.items()]
    wb = Workbook()
    sheet = None
    if len(wb.get_sheet_names()) == 0:
        sheet = wb.create_sheet("Sheet1")
    else:
        sheet = wb.active

    for row, r in enumerate(result):
        for col, v in enumerate(r):
            sheet.cell(column=col+1, row=row+1).value = v

    wb.save(outFile)
        
# TODO: update this function
"""
def promptExcelConcat():
    print('Please enter the name of the directory:')
    dir = input('>> ')
    while not os.path.exists(dir):
        print('That path does not exist. Try again:')
        dir = input('>> ')

    print('Enter the name of the output file:')
    output = input('>> ')
    while True:
        print('Press enter to accept or type a different path:')
        print('\t' + output)
        new_output = input('>> ')
        if new_output == '':
            break
        else:
            output = new_output
    
    concat_directory(dir, os.path.join(os.getcwd(), output),
                     keepFirstHeader=True);
"""
